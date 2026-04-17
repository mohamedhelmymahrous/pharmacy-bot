import os
import io
import re
import gc
import logging
import shutil
import tempfile
from difflib import SequenceMatcher

from telegram import Update
from telegram.ext import Application, MessageHandler, filters, ContextTypes
import pdfplumber
import openpyxl

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")

# ===========================================================================
# ===========================================================================

ITEM_RE = re.compile(r'^\d+\s+\d{3}-\d{5}-(.*?)\s+UOM:\s*(.+)$')
TOTAL4_RE = re.compile(
    r'^Total\s+([\d,]+(?:\.\d+)?)\s+([\d,]+(?:\.\d+)?)'
    r'\s+([\d,]+(?:\.\d+)?)\s+([\d,]+(?:\.\d+)?)$'
)
TOTAL3_RE = re.compile(
    r'^Total\s+([\d,]+(?:\.\d+)?)\s+([\d,]+(?:\.\d+)?)\s+([\d,]+(?:\.\d+)?)$'
)

def to_num(s):
    return float(s.replace(',', ''))

def extract_pdf_items(pdf_bytes):
    items = {}
    current = None

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for raw_line in text.split('\n'):
                line = raw_line.strip()

                m = ITEM_RE.match(line)
                if m:
                    current = {'name': m.group(1).strip().upper()}
                    continue

                m4 = TOTAL4_RE.match(line)
                if m4 and current:
                    current['bfw']      = to_num(m4.group(1))
                    current['received'] = to_num(m4.group(2))
                    current['issued']   = to_num(m4.group(3))
                    current['balance']  = to_num(m4.group(4))
                    items[current['name']] = current
                    current = None
                    continue

                m3 = TOTAL3_RE.match(line)
                if m3 and current:
                    current['bfw']      = 0.0
                    current['received'] = to_num(m3.group(1))
                    current['issued']   = to_num(m3.group(2))
                    current['balance']  = to_num(m3.group(3))
                    items[current['name']] = current
                    current = None

    gc.collect()
    return items


# ===========================================================================
# ===========================================================================

GENERIC = {
    'FILM', 'COATED', 'ORAL', 'EXTENDED', 'RELEASE', 'MODIFIED', 'COMP',
    'CHEWABLE', 'INFUSION', 'SACHET', 'POWDER', 'PATCH', 'EFFERVESCENT',
    'ENTERIC', 'SUBLINGUAL', 'TOPICAL', 'FORTE', 'PLUS', 'MINI', 'MICRO',
    'NANO', 'RETARD', 'DEPOT', 'LONG', 'SLOW', 'INSTANT', 'RAPID', 'SOFT',
    'HARD', 'GELATIN', 'SUPPOSITORY', 'ENEMA', 'PACK', 'STRIP',
}
FORMS = {
    'TABLET', 'TABLETS', 'TABS', 'CAPSULE', 'CAPSULES', 'CAPS',
    'SYRUP', 'SUSPENSION', 'DROPS', 'CREAM', 'OINTMENT',
    'INJECTION', 'AMPOULE', 'INHALER', 'SPRAY', 'LOTION',
    'GEL', 'SOLUTION', 'VIAL', 'CHEW',
}

def normalize(name):
    return re.sub(r'(\d+)\s+(MG|ML|MCG|IU|G)\b', r'\1\2', name.upper())

def key_number(name):
    m = re.search(r'(\d+(?:\.\d+)?)(?:MG|ML|MCG|IU|G)(?:/\d+(?:MG|ML))?', name)
    if m:
        return m.group(1)
    m = re.search(r'\b(\d+)\s*$', name.rstrip())
    if m:
        return m.group(1)
    return None

def meaningful_words(name):
    return [w for w in name.split()
            if len(w) > 3 and w not in GENERIC and w not in FORMS]

def form_words(name):
    return set(name.split()) & FORMS

def match_score(pdf_name, excel_name):
    pdf_n   = normalize(pdf_name)
    excel_n = normalize(excel_name)

    if pdf_n == excel_n:
        return 1.0

    emw = meaningful_words(excel_n)
    pmw = meaningful_words(pdf_n)

    base = (any(w in pdf_n for w in emw) or
            any(w in excel_n for w in pmw[:2]))
    if not base:
        return SequenceMatcher(None, pdf_n[:20], excel_n[:20]).ratio()

    score = 0.75

    pn = key_number(pdf_n)
    en = key_number(excel_n)
    if pn and en:
        score += 0.20 if pn == en else -0.40

    pf = form_words(pdf_n)
    ef = form_words(excel_n)
    if pf and ef:
        score += 0.10 if pf & ef else -0.40
    elif pf and not ef:
        ewc = len([w for w in excel_n.split() if w not in GENERIC])
        if not (ewc <= 2 and not en):
            score -= 0.10

    return max(0.0, min(score, 1.0))

def find_best_match(excel_name, pdf_items):
    best_score = 0.0
    best_item  = None
    for pdf_name, item in pdf_items.items():
        s = match_score(pdf_name, excel_name)
        if s > best_score:
            best_score = s
            best_item  = item
    return best_item if best_score >= 0.75 else None


# ===========================================================================
# ===========================================================================

def build_new_month(prev_excel_bytes, pdf_items):
    """
    Takes previous month Excel + new month PDF
    Returns new Excel where:
      E (balance)  = K from last month (remaining)
      F (received) = Received from PDF
      G, H         = empty
      I, J, K      = formulas stay as-is
      L:BU         = empty for daily manual entry
    """
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(prev_excel_bytes)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb['Sheet1']

        wb_data = openpyxl.load_workbook(tmp_path, data_only=True)
        ws_data = wb_data['Sheet1']

        excel_index = {}
        for row_idx in range(3, ws.max_row + 1):
            name_cell = ws_data.cell(row=row_idx, column=2).value
            if not name_cell or not str(name_cell).strip():
                continue
            k_val = ws_data.cell(row=row_idx, column=11).value or 0
            excel_index[row_idx] = {
                'name':    str(name_cell).strip(),
                'k_value': k_val
            }

        wb_data.close()

        matched   = 0
        unmatched = 0

        for row_idx, info in excel_index.items():
            excel_name = info['name']
            prev_remaining = info['k_value']

            ws.cell(row=row_idx, column=5).value = prev_remaining if prev_remaining else 0

            pdf_item = find_best_match(excel_name, pdf_items)

            if pdf_item:
                ws.cell(row=row_idx, column=6).value = (
                    int(pdf_item['received']) if pdf_item['received'] else None
                )
                matched += 1
            else:
                ws.cell(row=row_idx, column=6).value = None
                unmatched += 1

            ws.cell(row=row_idx, column=7).value = None
            ws.cell(row=row_idx, column=8).value = None

            for col in range(12, 74):
                cell = ws.cell(row=row_idx, column=col)
                if cell.value and not str(cell.value).startswith('='):
                    cell.value = None

        wb.save(tmp_path)
        wb.close()

        with open(tmp_path, 'rb') as f:
            result = f.read()

        return result, matched, unmatched

    finally:
        os.unlink(tmp_path)
        gc.collect()


# ===========================================================================
# ===========================================================================

pending = {}

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    msg     = update.message
    if not msg:
        return

    if msg.text:
        if msg.text.strip() == '/start':
            await msg.reply_text(
                "ahlan! :)\n\n"
                "3mly el2ati:\n"
                "1. eb3at excel el-shahr el-faat\n"
                "2. eb3at PDF el-shahr el-gedid\n\n"
                "msh mohim el-tartib ok"
            )
        else:
            await msg.reply_text(
                "eb3atli:\n"
                "1 - Excel el-shahr el-faat\n"
                "2 - Stock Card Report PDF"
            )
        return

    if msg.document:
        doc  = msg.document
        fname = (doc.file_name or "").lower()
        fext  = fname.split('.')[-1]

        file_obj = await context.bot.get_file(doc.file_id)
        data     = bytes(await file_obj.download_as_bytearray())

        cid = str(chat_id)
        if cid not in pending:
            pending[cid] = {}

        if fext == 'pdf':
            pending[cid]['pdf'] = data
            if 'excel' not in pending[cid]:
                await msg.reply_text("PDF ok - eb3at el-excel")

        elif fext in ('xlsx', 'xls'):
            pending[cid]['excel'] = data
            if 'pdf' not in pending[cid]:
                await msg.reply_text("Excel ok - eb3at el-PDF")

        else:
            await msg.reply_text("PDF aw Excel bas!")
            return

        if 'pdf' in pending[cid] and 'excel' in pending[cid]:
            files          = pending.pop(cid)
            pdf_bytes      = files['pdf']
            excel_bytes    = files['excel']

            await msg.reply_text("gari el-moa3ala...")

            try:
                pdf_items = extract_pdf_items(pdf_bytes)

                if not pdf_items:
                    await msg.reply_text(
                        "msh 2adr a2ra el-PDF\n"
                        "ta2akad enno Stock Card Report"
                    )
                    return

                result_bytes, matched, unmatched = build_new_month(
                    excel_bytes, pdf_items
                )

                await context.bot.send_document(
                    chat_id=chat_id,
                    document=io.BytesIO(result_bytes),
                    filename="gard_gedid.xlsx",
                    caption=(
                        "gard gahiz!\n\n"
                        "PDF items: " + str(len(pdf_items)) + "\n"
                        "mlab2: " + str(matched) + "\n"
                        "mafesh haraka: " + str(unmatched)
                    )
                )

            except Exception as e:
                logger.error(f"Error: {e}")
                await msg.reply_text("khata: " + str(e))

            finally:
                gc.collect()


def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(MessageHandler(filters.ALL, handle_message))
    logger.info("bot shaghal!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
