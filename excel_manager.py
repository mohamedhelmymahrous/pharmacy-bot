"""
excel_manager.py
----------------
Reads and updates inventory Excel file (منصرف_شهري.xlsx).

Excel structure (header on row 2):
  Col A: رقم 118     ← never touched
  Col B: كود         ← updated from PDF
  Col C: اسم الصنف   ← updated from PDF
  Col D: الوحدة      ← updated from PDF
  Col E: رصيد        ← BFW (opening balance)
  Col F: وارد 1      ← first received batch
  Col G: وارد 2      ← second received batch
  Col H: وارد 3      ← third received batch
  Col I: مجموع       ← formula: رصيد + كل الوارد
  Col J: منصرف       ← issued
  Col K: متبقى       ← formula: مجموع - منصرف
"""

import os
import logging
from typing import Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

EXCEL_PATH = os.environ.get("EXCEL_PATH", "منصرف_شهري.xlsx")

# Column positions (1-based)
COL_NUM    = 1   # رقم 118   ← never touched
COL_CODE   = 2   # كود
COL_NAME   = 3   # اسم الصنف
COL_UNIT   = 4   # الوحدة
COL_BFW    = 5   # رصيد
COL_REC1   = 6   # وارد 1
COL_REC2   = 7   # وارد 2
COL_REC3   = 8   # وارد 3
COL_TOTAL  = 9   # مجموع
COL_ISSUED = 10  # منصرف
COL_BAL    = 11  # متبقى

HEADER_ROW = 2   # row 1 is "دفتر 44", row 2 has column headers
DATA_START = 3   # data starts at row 3


# ---------------------------------------------------------------------------
# Load / Save
# ---------------------------------------------------------------------------

def load_excel() -> tuple[Workbook, Worksheet]:
    """Load existing Excel or create new one with correct headers."""
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        logger.info(f"Excel loaded: {EXCEL_PATH} ({ws.max_row - DATA_START + 1} items)")
    else:
        logger.warning("Excel not found — creating new file")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "دفتر 44"
        headers = ["رقم 118", "كود", "اسم الصنف ( E )", "الوحدة",
                   "رصيد", "وارد 1", "وارد 2", "وارد 3",
                   "مجموع", "منصرف", "متبقى"]
        for col, h in enumerate(headers, start=1):
            ws.cell(HEADER_ROW, col, h)
        wb.save(EXCEL_PATH)
    return wb, ws


def save_excel(wb: Workbook):
    """Save workbook to disk."""
    wb.save(EXCEL_PATH)
    logger.info(f"Excel saved → {EXCEL_PATH}")


# ---------------------------------------------------------------------------
# Read database
# ---------------------------------------------------------------------------

def sheet_to_db_list(ws: Worksheet) -> list[dict]:
    """
    Read all data rows and return as list of dicts.
    Used to build the PharmaMatcher database.
    """
    items = []
    for row in ws.iter_rows(min_row=DATA_START, values_only=False):
        name = row[COL_NAME - 1].value
        if not name:
            continue
        items.append({
            "id":       str(row[COL_NUM  - 1].value or ""),
            "code":     str(row[COL_CODE - 1].value or ""),
            "name":     str(name).strip(),
            "form":     str(row[COL_UNIT - 1].value or ""),
            "bfw":      _to_float(row[COL_BFW   - 1].value),
            "received": _to_float(row[COL_REC1  - 1].value) +
                        _to_float(row[COL_REC2  - 1].value) +
                        _to_float(row[COL_REC3  - 1].value),
            "issued":   _to_float(row[COL_ISSUED - 1].value),
            "balance":  _to_float(row[COL_BAL   - 1].value),
            "_row":     row[0].row,   # internal: actual Excel row number
        })
    return items


# ---------------------------------------------------------------------------
# Update
# ---------------------------------------------------------------------------

def update_excel(
    item: dict,
    match_type: str,
    matched_item: Optional[dict],
    ws: Worksheet,
) -> dict:
    """
    Update Excel based on match result.

    exact/fuzzy → find existing row by _row reference:
        - update name, code, unit from PDF
        - add received to next free وارد slot (1→2→3)
        - add issued to منصرف
        - recalculate مجموع and متبقى

    new → append new row at the end:
        - write name, code, unit
        - رقم 118 stays empty
        - write received in وارد 1, issued in منصرف
        - calculate مجموع and متبقى
    """
    received = _to_float(item.get("received", 0))
    issued   = _to_float(item.get("issued",   0))
    name     = str(item.get("name", "")).strip()
    code     = str(item.get("code", "")).strip()
    unit     = str(item.get("form", "")).strip()

    if match_type in ("exact", "fuzzy") and matched_item:
        row_num = matched_item.get("_row")
        if not row_num:
            logger.warning(f"No _row for matched item: {matched_item.get('name')}")
            return _append_new_row(ws, name, code, unit, received, issued)

        # Update name, code, unit from PDF
        ws.cell(row_num, COL_NAME).value = name
        ws.cell(row_num, COL_CODE).value = code
        ws.cell(row_num, COL_UNIT).value = unit

        # Add received to next free وارد slot
        if received > 0:
            if not ws.cell(row_num, COL_REC1).value:
                ws.cell(row_num, COL_REC1).value = received
            elif not ws.cell(row_num, COL_REC2).value:
                ws.cell(row_num, COL_REC2).value = received
            elif not ws.cell(row_num, COL_REC3).value:
                ws.cell(row_num, COL_REC3).value = received
            else:
                # All 3 slots full — add to وارد 3
                cur = _to_float(ws.cell(row_num, COL_REC3).value)
                ws.cell(row_num, COL_REC3).value = cur + received

        # Add issued
        if issued > 0:
            cur_issued = _to_float(ws.cell(row_num, COL_ISSUED).value)
            ws.cell(row_num, COL_ISSUED).value = cur_issued + issued

        # Recalculate مجموع and متبقى
        bfw   = _to_float(ws.cell(row_num, COL_BFW).value)
        rec1  = _to_float(ws.cell(row_num, COL_REC1).value)
        rec2  = _to_float(ws.cell(row_num, COL_REC2).value)
        rec3  = _to_float(ws.cell(row_num, COL_REC3).value)
        total = bfw + rec1 + rec2 + rec3
        iss   = _to_float(ws.cell(row_num, COL_ISSUED).value)

        ws.cell(row_num, COL_TOTAL).value = total
        ws.cell(row_num, COL_BAL).value   = total - iss

        logger.info(f"Updated row {row_num}: {name} → issued={iss} balance={total - iss}")
        return {
            "name": name, "code": code, "unit": unit,
            "bfw": bfw, "received": rec1 + rec2 + rec3,
            "issued": iss, "balance": total - iss,
        }

    # New item
    return _append_new_row(ws, name, code, unit, received, issued)


def _append_new_row(
    ws: Worksheet,
    name: str, code: str, unit: str,
    received: float, issued: float,
) -> dict:
    """Append a new item row. رقم 118 stays empty."""
    # Find first truly empty row after header
    next_row = ws.max_row + 1
    # Scan back in case there are empty rows at the bottom
    for r in range(ws.max_row, DATA_START - 1, -1):
        if any(ws.cell(r, c).value for c in range(1, 12)):
            next_row = r + 1
            break

    bfw     = 0.0
    total   = bfw + received
    balance = total - issued

    ws.cell(next_row, COL_NUM).value    = None        # رقم 118 — empty
    ws.cell(next_row, COL_CODE).value   = code
    ws.cell(next_row, COL_NAME).value   = name
    ws.cell(next_row, COL_UNIT).value   = unit
    ws.cell(next_row, COL_BFW).value    = bfw
    ws.cell(next_row, COL_REC1).value   = received if received > 0 else None
    ws.cell(next_row, COL_REC2).value   = None
    ws.cell(next_row, COL_REC3).value   = None
    ws.cell(next_row, COL_TOTAL).value  = total
    ws.cell(next_row, COL_ISSUED).value = issued if issued > 0 else None
    ws.cell(next_row, COL_BAL).value    = balance

    logger.info(f"New row {next_row}: {name} code={code}")
    return {
        "name": name, "code": code, "unit": unit,
        "bfw": bfw, "received": received,
        "issued": issued, "balance": balance,
    }


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _to_float(val) -> float:
    try:
        return float(val or 0)
    except (ValueError, TypeError):
        return 0.0
